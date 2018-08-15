import os
import time
from test_base import driver, slash
from openpyxl import load_workbook
from customer_search import actions
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains

__author__ = 'Nick Hartley'
# 5/29/2018


def verify_customers_imported(file):
    # list of customer numbers in the import file
    customer_numbers = read_import_file_customer_numbers(file)

    actions.customer_criteria_click('customer numbers input')

    for number in customer_numbers:
        ActionChains(driver).send_keys(number).send_keys(Keys.RETURN).perform()

    # time.sleep(2)

    actions.click('get search results button')

    time.sleep(3)

    count = driver.find_element_by_xpath('//*[@id="human_readable"]').text
    # 'count' is a string which starts with 'Found 14' e.g.
    # The following block checks if the number is a single or double digit number
    if count[7] == ' ':
        count = int(count[6])
    elif count[8] == ' ':
        count = int(count[6:8])
    assert count == len(customer_numbers)
    print('Customers imported successfully.')


def read_import_file_customer_numbers(file):
    path = '{0}{1}test_assets{1}{2}'.format(os.getcwd(), slash, file)
    print(path)
    wb = load_workbook(filename=path)
    ws = wb['Template']

    customer_numbers = []
    cell_data = ''
    row = 2
    while cell_data is not None:
        cell = 'A{}'.format(row)
        cell_data = ws[cell].value
        if cell_data is not None:
            # print(cell_data)
            # print(type(cell_data))
            customer_numbers.append(cell_data)
            row += 1
    print(customer_numbers)
    return customer_numbers


def verify_multi_zone_certificate(zones):
    screen_zones = []

    check = True
    x = 2  # Values start in tr[2]
    while check:
        try:
            screen_zone = driver.find_element_by_xpath(
                '//*[@id="search_results_grid"]/tbody/tr[2]/td[6]/table/tbody/tr[{}]/td[2]'.format(x)
            ).text

            screen_zones.append(screen_zone)
            x += 1
        except NoSuchElementException:
            check = False

    assert len(screen_zones) == len(zones), \
        'Incorrect number of exposure zones found. Expected: {}. Found {}.'.format(len(zones), len(screen_zones))
    print('Found {} exposure zones.'.format(len(screen_zones)))

    # Check each zone on the screen matches with an expected zone
    for screen_zone in screen_zones:
        assert any(screen_zone == zone for zone in zones), 'Unexpected exposure zone found: {}'.format(screen_zone)
        print('Found expected exposure zone: {}'.format(screen_zone))
